"""
https://stackoverflow.com/questions/71996770/not-supported-between-instances-of-tuple-and-int-error-in-python-seleniu
"""

if __name__ == '__main__':
    from openpyxl import Workbook, load_workbook
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time

    wb = load_workbook('data/data.xlsx')
    ws = wb.active

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_col=2, max_row=6, values_only=True)):

        driver = webdriver.Chrome()
        driver.get("https://www.google.com")

        driver.get("https://vehicleenquiry.service.gov.uk/")

        time.sleep(5)

        search = driver.find_element(By.ID, "wizard_vehicle_enquiry_capture_vrn_vrn")
        search.send_keys(str(row))
        search.send_keys(Keys.RETURN)

        try:
            main = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "summary-no-action"))
            )

            print(WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH,
                                                  "//dt[text()='Colour']/following::dd[1]"))).text
                  )

            ws.cell(row=row_index+2, column=2).value = str(driver.find_element(By.XPATH, "//dt[text()='Colour']/following::dd[1]").text)

            time.sleep(5)

        finally:
            wb.save('data.xlsx')
            driver.quit()